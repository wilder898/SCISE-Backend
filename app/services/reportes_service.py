import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Optional
from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.repositories.movimiento_repository import (
    get_resumen_movimientos,
    list_movimientos_historial,
    list_movimientos_historial_for_export,
)

EXPORT_HEADERS = [
    "Movimiento ID",
    "Tipo Movimiento",
    "Fecha Registro",
    "Equipo ID",
    "Serial",
    "Equipo Nombre",
    "Equipo Descripcion",
    "Estado",
    "Usuario Registro",
    "Estudiante",
    "Documento Estudiante",
]

EXPORT_COLUMN_WIDTHS = [16, 20, 22, 12, 22, 26, 36, 14, 24, 24, 22]
PDF_TABLE_COLUMNS = [
    ("ID", 8),
    ("TIPO", 8),
    ("FECHA", 19),
    ("SERIAL", 16),
    ("USUARIO", 20),
    ("ESTUDIANTE", 20),
    ("DOCUMENTO", 14),
]


def _normalizar_tipo(tipo: Optional[str], tipo_movimiento: Optional[str]) -> Optional[str]:
    value = (tipo_movimiento or tipo or "").strip().upper()
    if not value:
        return None
    if value.startswith("INGRESO"):
        return "INGRESO"
    if value.startswith("SALIDA"):
        return "SALIDA"
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="tipo debe ser INGRESO o SALIDA",
    )


def _normalizar_fechas(
    fecha: Optional[date],
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
    fecha_inicio: Optional[date],
    fecha_fin: Optional[date],
) -> tuple[Optional[date], Optional[date], Optional[date]]:
    fecha_desde_final = fecha_desde or fecha_inicio
    fecha_hasta_final = fecha_hasta or fecha_fin

    if fecha_desde_final and fecha_hasta_final and fecha_desde_final > fecha_hasta_final:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="fecha_desde no puede ser mayor que fecha_hasta",
        )

    return fecha, fecha_desde_final, fecha_hasta_final


def obtener_resumen_movimientos(
    db: Session,
    tipo: Optional[str] = None,
    tipo_movimiento: Optional[str] = None,
    fecha: Optional[date] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
) -> dict:
    tipo_normalizado = _normalizar_tipo(tipo=tipo, tipo_movimiento=tipo_movimiento)
    fecha_filtro, fecha_desde_filtro, fecha_hasta_filtro = _normalizar_fechas(
        fecha=fecha,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    return get_resumen_movimientos(
        db=db,
        tipo=tipo_normalizado,
        fecha=fecha_filtro,
        fecha_desde=fecha_desde_filtro,
        fecha_hasta=fecha_hasta_filtro,
    )


def listar_historial_movimientos(
    db: Session,
    tipo: Optional[str] = None,
    tipo_movimiento: Optional[str] = None,
    fecha: Optional[date] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    skip: Optional[int] = None,
    limit: Optional[int] = None,
    page: Optional[int] = None,
    page_size: Optional[int] = None,
) -> dict:
    tipo_normalizado = _normalizar_tipo(tipo=tipo, tipo_movimiento=tipo_movimiento)
    fecha_filtro, fecha_desde_filtro, fecha_hasta_filtro = _normalizar_fechas(
        fecha=fecha,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    effective_limit = limit if limit is not None else (page_size if page_size is not None else 5)
    effective_limit = max(1, min(int(effective_limit), 500))

    if skip is None:
        effective_page = max(1, int(page or 1))
        effective_skip = (effective_page - 1) * effective_limit
    else:
        effective_skip = max(0, int(skip))
        effective_page = (effective_skip // effective_limit) + 1

    registros, total = list_movimientos_historial(
        db=db,
        tipo=tipo_normalizado,
        fecha=fecha_filtro,
        fecha_desde=fecha_desde_filtro,
        fecha_hasta=fecha_hasta_filtro,
        skip=effective_skip,
        limit=effective_limit,
    )

    data = [
        {
            "id": int(registro.movimiento_id),
            "movimiento_id": int(registro.movimiento_id),
            "tipo_movimiento": registro.tipo_movimiento,
            "fecha_registro": registro.fecha_registro,
            "equipo_id": int(registro.equipo_id),
            "serial": registro.serial,
            "equipo_nombre": registro.equipo_nombre,
            "equipo_descripcion": registro.equipo_descripcion,
            "estado": registro.estado,
            "usuario_nombre": registro.usuario_nombre,
            "estudiante_nombre": registro.estudiante_nombre,
            "estudiante_documento": registro.estudiante_documento,
        }
        for registro in registros
    ]

    total_pages = max(1, (total + effective_limit - 1) // effective_limit)
    return {
        "data": data,
        "meta": {
            "total": total,
            "skip": effective_skip,
            "limit": effective_limit,
            "has_next": effective_skip + len(data) < total,
        },
        "total": total,
        "page": effective_page,
        "total_pages": total_pages,
    }


def obtener_resumen_dashboard(
    db: Session,
    tipo: Optional[str] = None,
    tipo_movimiento: Optional[str] = None,
    fecha: Optional[date] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
) -> dict:
    return obtener_resumen_movimientos(
        db=db,
        tipo=tipo,
        tipo_movimiento=tipo_movimiento,
        fecha=fecha,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )


def listar_historial_reciente_dashboard(
    db: Session,
    tipo: Optional[str] = None,
    tipo_movimiento: Optional[str] = None,
    fecha: Optional[date] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    page: int = 1,
    limit: int = 5,
) -> dict:
    safe_page = max(1, int(page or 1))
    safe_limit = max(1, min(int(limit or 5), 5))
    return listar_historial_movimientos(
        db=db,
        tipo=tipo,
        tipo_movimiento=tipo_movimiento,
        fecha=fecha,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        page=safe_page,
        page_size=safe_limit,
    )


def exportar_historial_movimientos_csv(
    db: Session,
    tipo: Optional[str] = None,
    tipo_movimiento: Optional[str] = None,
    fecha: Optional[date] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
) -> str:
    tipo_normalizado = _normalizar_tipo(tipo=tipo, tipo_movimiento=tipo_movimiento)
    fecha_filtro, fecha_desde_filtro, fecha_hasta_filtro = _normalizar_fechas(
        fecha=fecha,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    registros = list_movimientos_historial_for_export(
        db=db,
        tipo=tipo_normalizado,
        fecha=fecha_filtro,
        fecha_desde=fecha_desde_filtro,
        fecha_hasta=fecha_hasta_filtro,
    )

    output = StringIO(newline="")
    writer = csv.writer(output, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(EXPORT_HEADERS)
    for registro in registros:
        writer.writerow(_serialize_registro_for_export(registro))

    # Excel en configuraciones regionales es-CO suele usar ';' como separador.
    # 'sep=;' fuerza la separacion por columnas al abrir el archivo.
    csv_content = output.getvalue()
    return "\ufeffsep=;\n" + csv_content


def exportar_historial_movimientos_xlsx(
    db: Session,
    tipo: Optional[str] = None,
    tipo_movimiento: Optional[str] = None,
    fecha: Optional[date] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
) -> tuple[str, bytes]:
    tipo_normalizado = _normalizar_tipo(tipo=tipo, tipo_movimiento=tipo_movimiento)
    fecha_filtro, fecha_desde_filtro, fecha_hasta_filtro = _normalizar_fechas(
        fecha=fecha,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    registros = list_movimientos_historial_for_export(
        db=db,
        tipo=tipo_normalizado,
        fecha=fecha_filtro,
        fecha_desde=fecha_desde_filtro,
        fecha_hasta=fecha_hasta_filtro,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Historial Movimientos"
    worksheet.append(EXPORT_HEADERS)

    for registro in registros:
        worksheet.append(_serialize_registro_for_export(registro))

    header_fill = PatternFill(fill_type="solid", fgColor="166534")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index, _ in enumerate(EXPORT_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.column_dimensions[get_column_letter(column_index)].width = EXPORT_COLUMN_WIDTHS[column_index - 1]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_HEADERS))}{max(1, worksheet.max_row)}"
    worksheet.sheet_view.zoomScale = 110

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return "reporte-movimientos.xlsx", output.read()


def exportar_historial_movimientos_pdf(
    db: Session,
    tipo: Optional[str] = None,
    tipo_movimiento: Optional[str] = None,
    fecha: Optional[date] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
) -> tuple[str, bytes]:
    tipo_normalizado = _normalizar_tipo(tipo=tipo, tipo_movimiento=tipo_movimiento)
    fecha_filtro, fecha_desde_filtro, fecha_hasta_filtro = _normalizar_fechas(
        fecha=fecha,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    registros = list_movimientos_historial_for_export(
        db=db,
        tipo=tipo_normalizado,
        fecha=fecha_filtro,
        fecha_desde=fecha_desde_filtro,
        fecha_hasta=fecha_hasta_filtro,
    )

    filtros = []
    if tipo_normalizado:
        filtros.append(f"tipo={tipo_normalizado}")
    if fecha_filtro:
        filtros.append(f"fecha={fecha_filtro.isoformat()}")
    if fecha_desde_filtro:
        filtros.append(f"fecha_desde={fecha_desde_filtro.isoformat()}")
    if fecha_hasta_filtro:
        filtros.append(f"fecha_hasta={fecha_hasta_filtro.isoformat()}")

    table_header = _build_pdf_table_header()
    table_separator = _build_pdf_table_separator()
    table_rows = [_build_pdf_table_row(registro) for registro in registros]

    generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    filtros_texto = ", ".join(filtros) if filtros else "sin filtros"

    max_body_rows_per_page = 34
    total_pages = max(1, ((len(table_rows) or 1) + max_body_rows_per_page - 1) // max_body_rows_per_page)
    pages_lines: list[list[str]] = []

    for page_index in range(total_pages):
        start = page_index * max_body_rows_per_page
        end = start + max_body_rows_per_page
        page_rows = table_rows[start:end]

        lines: list[str] = []
        lines.extend(_wrap_pdf_line("SCISE - Reporte de Movimientos", 110))
        lines.extend(_wrap_pdf_line(f"Generado: {generated_at}", 110))
        lines.extend(_wrap_pdf_line(f"Filtros: {filtros_texto}", 110))
        lines.extend(_wrap_pdf_line(f"Total registros: {len(registros)}", 110))
        lines.extend(_wrap_pdf_line(f"Pagina {page_index + 1} de {total_pages}", 110))
        lines.append("")

        if not registros:
            lines.extend(_wrap_pdf_line("No hay movimientos para los filtros seleccionados.", 110))
        else:
            lines.append(table_header)
            lines.append(table_separator)
            lines.extend(page_rows)
            lines.append(table_separator)

        pages_lines.append(lines)

    return "reporte-movimientos.pdf", _build_multipage_pdf(pages_lines)


def _serialize_registro_for_export(registro) -> list[str | int]:
    return [
        int(registro.movimiento_id),
        registro.tipo_movimiento,
        registro.fecha_registro.strftime("%Y-%m-%d %H:%M:%S") if registro.fecha_registro else "",
        int(registro.equipo_id),
        registro.serial or "",
        registro.equipo_nombre or "",
        registro.equipo_descripcion or "",
        registro.estado or "",
        registro.usuario_nombre or "",
        registro.estudiante_nombre or "",
        registro.estudiante_documento or "",
    ]


def _escape_pdf_text(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _fit_pdf_column(text: str, width: int) -> str:
    normalized = str(text or "").replace("\n", " ").strip()
    if len(normalized) > width:
        if width <= 3:
            normalized = normalized[:width]
        else:
            normalized = normalized[: width - 3] + "..."
    return normalized.ljust(width)


def _build_pdf_table_header() -> str:
    return " | ".join(_fit_pdf_column(title, width) for title, width in PDF_TABLE_COLUMNS)


def _build_pdf_table_separator() -> str:
    return "-+-".join("-" * width for _, width in PDF_TABLE_COLUMNS)


def _build_pdf_table_row(registro) -> str:
    fecha_texto = (
        registro.fecha_registro.strftime("%Y-%m-%d %H:%M:%S")
        if registro.fecha_registro
        else "-"
    )
    values = [
        int(registro.movimiento_id),
        registro.tipo_movimiento or "",
        fecha_texto,
        registro.serial or "",
        registro.usuario_nombre or "",
        registro.estudiante_nombre or "",
        registro.estudiante_documento or "",
    ]
    formatted = [
        _fit_pdf_column(str(value), width)
        for value, (_, width) in zip(values, PDF_TABLE_COLUMNS)
    ]
    return " | ".join(formatted)


def _wrap_pdf_line(text: str, max_width: int) -> list[str]:
    value = str(text or "")
    if len(value) <= max_width:
        return [value]

    lines: list[str] = []
    current = value
    while len(current) > max_width:
        split_at = current.rfind(" ", 0, max_width)
        if split_at < 1:
            split_at = max_width
        lines.append(current[:split_at].strip())
        current = current[split_at:].strip()
    if current:
        lines.append(current)
    return lines or [""]


def _build_pdf_page_stream(lines: list[str]) -> bytes:
    safe_lines = [line if isinstance(line, str) else str(line) for line in lines]
    max_lines = 54
    trimmed_lines = safe_lines[:max_lines]

    content_parts = ["BT", "/F1 9 Tf", "35 810 Td"]
    for idx, line in enumerate(trimmed_lines):
        if idx > 0:
            content_parts.append("0 -14 Td")
        content_parts.append(f"({_escape_pdf_text(line)}) Tj")
    content_parts.append("ET")

    content_text = "\n".join(content_parts) + "\n"
    return content_text.encode("latin-1", errors="replace")


def _build_multipage_pdf(pages_lines: list[list[str]]) -> bytes:
    normalized_pages = pages_lines or [["SCISE - Reporte sin contenido"]]

    page_object_ids: list[int] = []
    content_object_ids: list[int] = []
    objects: dict[int, bytes] = {
        1: b"<< /Type /Catalog /Pages 2 0 R >>",
        2: b"",  # filled after creating page objects
        3: b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>",
    }

    next_object_id = 4
    for page_lines in normalized_pages:
        content_bytes = _build_pdf_page_stream(page_lines)
        content_object_id = next_object_id
        page_object_id = next_object_id + 1
        next_object_id += 2

        objects[content_object_id] = (
            f"<< /Length {len(content_bytes)} >>\n".encode("ascii")
            + b"stream\n"
            + content_bytes
            + b"endstream"
        )
        objects[page_object_id] = (
            b"<< /Type /Page /Parent 2 0 R "
            b"/MediaBox [0 0 595 842] "
            b"/Resources << /Font << /F1 3 0 R >> >> "
            + f"/Contents {content_object_id} 0 R >>".encode("ascii")
        )

        content_object_ids.append(content_object_id)
        page_object_ids.append(page_object_id)

    kids = " ".join(f"{page_id} 0 R" for page_id in page_object_ids)
    objects[2] = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_object_ids)} >>".encode("ascii")

    max_object_id = max(objects.keys())
    pdf = b"%PDF-1.4\n"
    offsets = [0] * (max_object_id + 1)

    for object_id in range(1, max_object_id + 1):
        offsets[object_id] = len(pdf)
        pdf += f"{object_id} 0 obj\n".encode("ascii")
        pdf += objects[object_id]
        pdf += b"\nendobj\n"

    xref_offset = len(pdf)
    pdf += f"xref\n0 {max_object_id + 1}\n".encode("ascii")
    pdf += b"0000000000 65535 f \n"
    for object_id in range(1, max_object_id + 1):
        pdf += f"{offsets[object_id]:010d} 00000 n \n".encode("ascii")

    pdf += f"trailer\n<< /Size {max_object_id + 1} /Root 1 0 R >>\n".encode("ascii")
    pdf += b"startxref\n"
    pdf += f"{xref_offset}\n".encode("ascii")
    pdf += b"%%EOF\n"
    return pdf
